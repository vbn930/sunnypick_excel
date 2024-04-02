#python version 3.10.11
#used lib : pandas

#pyinstaller -n "써니픽 정산서 엑셀 제작 프로그램" --clean --onefile main.py
import pandas as pd
import shutil
import os
from dataclasses import dataclass
from datetime import datetime
import time
import openpyxl
from openpyxl.styles import Font, Border, Side

@dataclass
class OrderData:
    order_date: str
    shop_name: str
    dealer_name: str
    brand_name: str
    item_name: str #상품명 - 옵션명(송장출력명)
    quantity: int
    order_num: str
    delivery_name: str
    delivery_num: str
    reciever_name: str
    reciever_phone: str
    selling_price: int
    supply_price: int #매출단가 - 공급단가
    total_supply_price: int #공급단가 총계
    delivery_fee: int
    extra_delivery_fee: int

def get_duration_from_csv():

    print("예시와 같은 포맷으로 정산 기간을 설정 해주세요.")
    print("예시 : 2024-03-31")
    start_date = ""
    end_date = ""
    while(True):
        start_date = input("정산 시작 날짜를 입력 해주세요 : ")
        if len(start_date) == 10:
            break
        else:
            print("입력 포맷이 잘못되었습니다. 다시 입력 해주세요.")
    
    while(True):
        end_date = input("정산 종료 날짜를 입력 해주세요 : ")
        if len(end_date) == 10:
            break
        else:
            print("입력 포맷이 잘못되었습니다. 다시 입력 해주세요.")
    return start_date, end_date

def get_spilted_list(org_data):
    dst_datas = []
    dst_data = []

    for i in range(len(org_data)):
        if i == 0:
            dst_data.append(org_data[i])
            if len(org_data) == 1:
                dst_datas.append(dst_data)
                break
        else:
            if org_data[i-1].shop_name != org_data[i].shop_name:
                dst_datas.append(dst_data)
                dst_data = []
            dst_data.append(org_data[i])
        if i == len(org_data) - 1:
            dst_datas.append(dst_data)
    return dst_datas

def get_input_data(file_path, start_date, end_date):
    order_datas = []
    datas = pd.read_excel(file_path)
    datas = datas.fillna(0)
    for idx, row in datas.iterrows():
        date = row["날짜"]
        if start_date <= date and date <= end_date:
            delivery_fee = row["배송비"]
            extra_delivery_fee = row["추가배송비"]
            if delivery_fee == "-" or delivery_fee == "":
                delivery_fee = 0
            if extra_delivery_fee == "-" or extra_delivery_fee == "":
                extra_delivery_fee = 0
            order_data = OrderData(row["날짜"], row["쇼핑몰명"], row["사업자명"], row["브랜드"], row["상품명"], row["수량"], 
                row["주문번호"], row["택배사"], row["운송장번호"], row["수취인명"], row["수취인 휴대폰"], row["판매가"], row["매출단가"], row["매출총계"], delivery_fee, extra_delivery_fee)
            order_datas.append(order_data)
    order_datas.sort(key=lambda x: x.shop_name)
    return  get_spilted_list(order_datas)

def get_shop_info(shop_data, date_duration, save_path):
    workbook = openpyxl.load_workbook("./template/output_template.xlsx", data_only=True)
    overview_sheet = workbook["요약"]
    detail_sheet = workbook["상세"]

    #요약 페이지 제목 설정
    overview_sheet.cell(row=1, column=1).value = f"주식회사 써니픽 정산서-{shop_data[0].shop_name}"

    shop_data.sort(key=lambda x: x.delivery_num)

    #운송장 번호가 일치하는 주문은 배송비 한번만 부과
    for i in range(len(shop_data)):
        if i != 0:
            if shop_data[i - 1].delivery_num == shop_data[i].delivery_num:
                shop_data[i].delivery_fee = 0
                shop_data[i].extra_delivery_fee = 0
    
    #브랜드 별로 각각 금액과 수량 계산
    shop_data.sort(key=lambda x: x.brand_name)
    total_price = 0
    total_delivery_fee = 0
    total_extra_delivery_fee = 0
    for data in shop_data:
        total_price += data.total_supply_price
        total_delivery_fee += data.delivery_fee
        total_extra_delivery_fee += data.extra_delivery_fee

    #주문 총계 요약 입력
    overview_sheet.cell(row=2, column=2).value = date_duration
    overview_sheet.cell(row=3, column=2).value = total_price
    overview_sheet.cell(row=4, column=2).value = total_delivery_fee
    overview_sheet.cell(row=5, column=2).value = total_extra_delivery_fee
    overview_sheet.cell(row=6, column=2).value = total_price + total_delivery_fee + total_extra_delivery_fee

    for i in range(4):
        overview_sheet.cell(row=i+3, column=2).number_format = "#,##0"
    overview_sheet.cell(row=6, column=2).font = Font(color="FF0000", bold=True)

    brand_data = get_brand_info(shop_data)

    last_idx = 9
    for brand in brand_data:
        brand.sort(key=lambda x: x.item_name)
        item_info_list = get_item_info(brand)

        #요약 페이지 정보 입력
        #row 는 9부터 시작
        for i in range(len(item_info_list)):
            for j in range(len(item_info_list[i])):
                overview_sheet.cell(row=(i+last_idx), column=(j+1)).value = item_info_list[i][j]
                overview_sheet.cell(row=(i+last_idx), column=(j+1)).font = Font(size=11)
                #엑셀 칸 포맷 변경
                if j == 5 or j == 6:
                    overview_sheet.cell(row=(i+last_idx), column=(j+1)).number_format = "#,##0"
            
                #엑셀 테두리 설정
                if j == (len(item_info_list[i]) - 1):
                    overview_sheet.cell(row=(i+last_idx), column=(j+1)).border = Border(right=Side(border_style='thin', color='000000'))
                if i == (len(item_info_list) - 1):
                    overview_sheet.cell(row=(i+last_idx), column=(j+1)).border = Border(bottom=Side(border_style='thin', color='000000'))
                if i == (len(item_info_list) - 1) and j == (len(item_info_list[i]) - 1):
                    overview_sheet.cell(row=(i+last_idx), column=(j+1)).border = Border(right=Side(border_style='thin', color='000000'), 
                                                                                        bottom=Side(border_style='thin', color='000000'))
        last_idx += len(item_info_list)

    shop_data = sorted(shop_data, key=lambda x : (x.order_date, x.brand_name, x.item_name))
    #상세 주문 정보 입력
    for i in range(len(shop_data)):
        #row 2 부터 시작
        detail_info = [shop_data[i].order_date, shop_data[i].shop_name, shop_data[i].dealer_name, shop_data[i].brand_name, shop_data[i].item_name, shop_data[i].quantity,
                       shop_data[i].order_num, shop_data[i].delivery_name, str(shop_data[i].delivery_num), shop_data[i].reciever_name, shop_data[i].reciever_phone,
                       shop_data[i].selling_price, shop_data[i].supply_price, shop_data[i].total_supply_price, shop_data[i].delivery_fee, shop_data[i].extra_delivery_fee]
        for j in range(len(detail_info)):
            #엑셀 칸 포맷 변경
            if j == 0:
                detail_sheet.cell(row=(i+2), column=(j+1)).number_format = "m월 d일 ;@"
            if j == 11 or j == 12 or j == 13 or j == 14 or j == 15:
                detail_sheet.cell(row=(i+2), column=(j+1)).number_format = "#,##0"

            detail_sheet.cell(row=(i+2), column=(j+1)).value = detail_info[j]
            detail_sheet.cell(row=(i+2), column=(j+1)).border = Border(top=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), 
                                                                       bottom=Side(border_style='thin', color='000000'),left=Side(border_style='thin', color='000000'))

    file_name = f"써니픽 정산서-{shop_data[0].shop_name}_{date_duration}.xlsx"
    print(f"엑셀 파일 \'{file_name}\'을 생성하였습니다!")
    workbook.save(f"{save_path}/{file_name}")

def get_brand_info(shop_data):
    org_data = shop_data
    dst_datas = []
    dst_data = []

    for i in range(len(org_data)):
        if i == 0:
            dst_data.append(org_data[i])
            if len(org_data) == 1:
                dst_datas.append(dst_data)
                break
        else:
            if org_data[i-1].brand_name != org_data[i].brand_name:
                dst_datas.append(dst_data)
                dst_data = []
            dst_data.append(org_data[i])
        if i == len(org_data) - 1:
            dst_datas.append(dst_data)
    
    brand_datas = dst_datas
    return brand_datas
    
def get_item_info(brand_data):
    org_data = brand_data
    dst_datas = []
    dst_data = []
    for i in range(len(org_data)):
        if i == 0:
            dst_data.append(org_data[i])
            if len(org_data) == 1:
                dst_datas.append(dst_data)
                break
        else:
            if org_data[i-1].item_name != org_data[i].item_name:
                dst_datas.append(dst_data)
                dst_data = []
            dst_data.append(org_data[i])
        if i == len(org_data) - 1:
            dst_datas.append(dst_data)
    item_datas = dst_datas
    brand_name = brand_data[0].brand_name
    item_info_list = []
    item_delivery_cnt = 0
    total_delivery_fee = 0
    item_extra_delivery_cnt = 0
    total_extra_delivery_fee = 0
    
    for item_data in  item_datas:
        item_info = ["","","","","","",""] #last idx = 6
        item_cnt = 0
        total_item_price = 0

        if item_data == item_datas[0]:
            item_info[0] = item_data[0].shop_name
            item_info[1] = item_data[0].dealer_name
        item_info[2] = item_data[0].brand_name
        item_info[3] = item_data[0].item_name
        for item in item_data:
            item_cnt += item.quantity
            total_item_price += item.total_supply_price
            if item.delivery_fee > 0:
                item_delivery_cnt += 1
                total_delivery_fee += item.delivery_fee
            if item.extra_delivery_fee > 0:
                item_extra_delivery_cnt += 1
                total_extra_delivery_fee += item.extra_delivery_fee
        item_info[4] = item_cnt
        item_info[5] = item_data[0].supply_price
        item_info[6] = total_item_price
        item_info_list.append(item_info)
    item_info_list.append(["","","",f"{brand_name} 배송비",item_delivery_cnt,"",total_delivery_fee])

    if total_extra_delivery_fee > 0:
        item_info_list.append(["","","",f"{brand_name} 추가배송비",item_extra_delivery_cnt,"",total_extra_delivery_fee])
    
    return item_info_list

def get_file_names():
    file_list = os.listdir("./input")
    file_names = [file for file in file_list if file.endswith(".xlsx")]
    print("input 폴더에서 아래의 파일을 발견하였습니다!")
    print(file_names, sep=", ")
    return file_names

def main():
    os.makedirs(f"./output", exist_ok=True)
    start_date = ""
    end_date = ""
    while(True):
        start_date, end_date = get_duration_from_csv()
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        if start_date > end_date:
            print("정산 시작 날짜가 정산 종료 날짜보다 늦습니다. 다시 입력해주세요.")
        else:
            break

    date_duration = start_date.strftime("%Y.%m.%d") + "-" + end_date.strftime("%m.%d")

    file_names = get_file_names()
    for file_name in file_names:
        os.makedirs(f"./output/{file_name[:-4]}", exist_ok=True)
        save_path = f"./output/{file_name[:-4]}"
        print(f"파일 \'{file_name}\' 의 작업을 시작합니다.")
        shop_datas = get_input_data(f"./input/{file_name}", start_date, end_date)
        for shop_data in shop_datas:
            get_shop_info(shop_data, date_duration, save_path)
    

if __name__ == "__main__":
    try:
        print("-써니픽 정산서 제작 프로그램-")
        main()
    except Exception as e:
        print(f"다음과 같은 오류가 발생하여 프로그램을 종료합니다. : {e}")
    finally:
        end = input("프로그램을 종료하려면 엔터키를 눌러주세요.")